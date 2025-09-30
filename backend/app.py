# backend/app.py
import io, os, re, json, base64, zipfile, asyncio, unicodedata
from typing import List, Optional, Dict, Any, Tuple
from urllib.parse import urlparse, parse_qs, quote
from uuid import uuid4
from datetime import timedelta

import httpx
from PIL import Image, UnidentifiedImageError, ImageFile
from fastapi import FastAPI, UploadFile, File, HTTPException, Request, Depends
from fastapi.responses import StreamingResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired
from openpyxl import load_workbook

# Pillow safety
ImageFile.LOAD_TRUNCATED_IMAGES = True

# -----------------------------------------------------------------------------
# Config (ENV)
# -----------------------------------------------------------------------------
SECRET_KEY = os.getenv("SECRET_KEY", "change-me")
SESSION_MAX_AGE = int(os.getenv("SESSION_MAX_AGE", "86400"))  # 24h
# Initial allowlist (admin can add/remove at runtime via API)
ALLOWED_EMAILS = {
    e.strip().lower()
    for e in os.getenv("ALLOWED_EMAILS", "").split(",")
    if e.strip()
}
# Admin token for allow/revoke endpoints (simple bearer)
ADMIN_TOKEN = os.getenv("ADMIN_TOKEN", "set-an-admin-token")

# Optional CORS (harmless if same-origin single container)
ALLOWED = os.getenv("CORS_ORIGIN", "*")
origins = [o.strip() for o in ALLOWED.split(",") if o.strip()]

# Browser-y UA helps some CDNs
DEFAULT_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0 Safari/537.36"
    )
}

# -----------------------------------------------------------------------------
# App + session serializer + in-memory stores
# -----------------------------------------------------------------------------
app = FastAPI(title="XLSX Image Zipper (Email-gated)")

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    max_age=3600,
)

# Signed cookie (no DB)
serializer = URLSafeTimedSerializer(SECRET_KEY, salt="session-v1")

# One active session per email (in-memory)
ACTIVE_SESSIONS: Dict[str, str] = {}  # email -> sid
# Dynamic allowlist (in-memory, seeded from env)
DYN_ALLOWED: Dict[str, bool] = {e: True for e in ALLOWED_EMAILS}

# -----------------------------------------------------------------------------
# Utilities
# -----------------------------------------------------------------------------
def slugify(text: str, max_len: int = 70) -> str:
    text = unicodedata.normalize("NFKD", str(text))
    text = text.encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^A-Za-z0-9._ -]+", "_", text)
    text = re.sub(r"\s+", "_", text)
    text = re.sub(r"_+", "_", text).strip("._ ")
    return (text or "untitled")[:max_len]

def ascii_fallback_filename(s: str) -> str:
    norm = unicodedata.normalize("NFKD", s)
    ascii_only = norm.encode("ascii", "ignore").decode("ascii")
    ascii_only = re.sub(r"[^A-Za-z0-9._ -]+", "_", ascii_only).strip()
    if not ascii_only:
        ascii_only = "images.zip"
    if not ascii_only.lower().endswith(".zip"):
        ascii_only += ".zip"
    return ascii_only

def extract_url_from_cell(cell) -> str:
    try:
        if getattr(cell, "hyperlink", None) and getattr(cell.hyperlink, "target", None):
            return str(cell.hyperlink.target).strip()
    except Exception:
        pass
    v = cell.value
    if not isinstance(v, str):
        return ""
    s = v.strip()
    if not s:
        return ""
    m = re.search(r'HYPERLINK\(\s*"([^"]+)"', s, re.I)
    if m:
        return m.group(1).strip()
    if s.startswith(("http://", "https://")):
        return s
    return ""

def strip_tracking_params(url: str) -> str:
    try:
        p = urlparse(url)
        if not p.query:
            return url
        qs = parse_qs(p.query)
        for k in ["utm_source", "utm_medium", "utm_campaign", "utm_term", "utm_content", "fbclid", "gclid"]:
            qs.pop(k, None)
        new_q = "&".join(f"{k}={v[0]}" for k, v in qs.items())
        return p._replace(query=new_q).geturl()
    except Exception:
        return url

def drive_fetch_plan(parsed, original_qs) -> List[str]:
    candidates: List[str] = [parsed.geturl()]
    file_id = None
    if parsed.path.startswith("/file/d/"):
        parts = parsed.path.split("/")
        if len(parts) > 3:
            file_id = parts[3]
    elif parsed.path.startswith("/open") and "id" in original_qs:
        file_id = original_qs["id"][0]
    elif parsed.path.startswith("/uc") and "id" in original_qs:
        file_id = original_qs["id"][0]
    if file_id:
        rk = original_qs.get("resourcekey", [None])[0]
        q = f"export=download&id={file_id}"
        if rk:
            q += f"&resourcekey={rk}"
        candidates.append(f"https://drive.google.com/uc?{q}")
    seen, uniq = set(), []
    for u in candidates:
        if u not in seen:
            uniq.append(u); seen.add(u)
    return uniq

def build_fetch_plan(raw_url: str) -> Tuple[str, List[str]]:
    display = raw_url.strip()
    if not display:
        return "", []
    try:
        p = urlparse(display)
        host = (p.netloc or "").lower()
        qs = parse_qs(p.query or "")
        if "drive.google.com" in host:
            return display, drive_fetch_plan(p, qs)
        if "dropbox.com" in host:
            cand = p
            if "dl=" in (p.query or ""):
                q = re.sub(r"dl=\d", "dl=1", p.query)
                cand = p._replace(query=q)
            return display, [display, cand.geturl()]
        stripped = strip_tracking_params(display)
        if stripped != display:
            return display, [display, stripped]
        return display, [display]
    except Exception:
        return display, [display]

async def fetch_one(client: httpx.AsyncClient, urls: List[str], sem: asyncio.Semaphore) -> bytes | None:
    for u in urls:
        async with sem:
            try:
                r = await client.get(u, timeout=45.0, follow_redirects=True, headers=DEFAULT_HEADERS)
                r.raise_for_status()
                return r.content
            except Exception:
                continue
    return None

def find_final_col(ws) -> Optional[Tuple[int, int]]:
    max_r = min(ws.max_row or 1, 500)
    for r in range(1, max_r + 1):
        for row in ws.iter_rows(min_row=r, max_row=r, values_only=False):
            for cell in row:
                val = cell.value
                if isinstance(val, str) and val.strip().lower().startswith("final"):
                    return (r, cell.column)
    return None

# -----------------------------------------------------------------------------
# Session helpers
# -----------------------------------------------------------------------------
def _load_session(request: Request) -> Optional[Dict[str, str]]:
    token = request.cookies.get("session")
    if not token:
        return None
    try:
        data = serializer.loads(token, max_age=SESSION_MAX_AGE)
        email = (data.get("email") or "").lower()
        sid = data.get("sid")
        if not email or not sid:
            return None
        # enforce one active session per email
        if ACTIVE_SESSIONS.get(email) != sid:
            return None
        return {"email": email, "sid": sid}
    except (BadSignature, SignatureExpired):
        return None

def require_user(request: Request) -> str:
    sess = _load_session(request)
    if not sess:
        raise HTTPException(status_code=401, detail="Not authenticated")
    return sess["email"]

# -----------------------------------------------------------------------------
# Auth endpoints (email only)
# -----------------------------------------------------------------------------
@app.post("/auth/login")
async def email_login(request: Request):
    """
    Body JSON: { "email": "user@example.com" }
    - 200 if allowed and not active elsewhere (sets cookie)
    - 403 if not allowed
    - 409 if already active on another device
    """
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    email = (payload.get("email") or "").strip().lower()
    if not email or "@" not in email:
        return JSONResponse({"error": "Enter a valid email."}, status_code=400)

    # allowed?
    allowed = DYN_ALLOWED.get(email, False)
    if not allowed:
        return JSONResponse({"error": "Email not authorized. Contact admin."}, status_code=403)

    # already active elsewhere?
    if email in ACTIVE_SESSIONS:
        return JSONResponse({"error": "Session already active on another device."}, status_code=409)

    # create session
    sid = uuid4().hex
    ACTIVE_SESSIONS[email] = sid
    cookie_val = serializer.dumps({"email": email, "sid": sid})

    resp = JSONResponse({"ok": True, "email": email})
    resp.set_cookie(
        "session",
        cookie_val,
        max_age=SESSION_MAX_AGE,
        httponly=True,
        secure=True,
        samesite="lax",
        path="/",
    )
    return resp

@app.post("/auth/logout")
async def email_logout(request: Request):
    sess = _load_session(request)
    resp = JSONResponse({"ok": True})
    if sess:
        if ACTIVE_SESSIONS.get(sess["email"]) == sess["sid"]:
            ACTIVE_SESSIONS.pop(sess["email"], None)
    resp.delete_cookie("session", path="/")
    return resp

@app.get("/api/me")
def me(request: Request):
    sess = _load_session(request)
    if not sess:
        raise HTTPException(401, "Not authenticated")
    return {"email": sess["email"]}

# -----------------------------------------------------------------------------
# Admin endpoints (in-memory allowlist; simple bearer token)
# -----------------------------------------------------------------------------
def _require_admin(request: Request):
    auth = request.headers.get("Authorization", "")
    if not auth.startswith("Bearer "):
        raise HTTPException(401, "Missing bearer token")
    token = auth.split(" ", 1)[1].strip()
    if token != ADMIN_TOKEN:
        raise HTTPException(403, "Invalid token")

@app.post("/admin/allow")
async def admin_allow(request: Request):
    _require_admin(request)
    body = await request.json()
    email = (body.get("email") or "").strip().lower()
    if not email or "@" not in email:
        raise HTTPException(400, "Valid email required")
    DYN_ALLOWED[email] = True
    return {"ok": True, "allowed": sorted(DYN_ALLOWED.keys())}

@app.post("/admin/revoke")
async def admin_revoke(request: Request):
    _require_admin(request)
    body = await request.json()
    email = (body.get("email") or "").strip().lower()
    if not email or "@" not in email:
        raise HTTPException(400, "Valid email required")
    DYN_ALLOWED.pop(email, None)
    # if currently active, force logout
    if email in ACTIVE_SESSIONS:
        ACTIVE_SESSIONS.pop(email, None)
    return {"ok": True, "allowed": sorted(DYN_ALLOWED.keys())}

@app.get("/admin/state")
def admin_state(request: Request):
    _require_admin(request)
    return {
        "allowed": sorted(DYN_ALLOWED.keys()),
        "active": list(ACTIVE_SESSIONS.keys()),
    }

# -----------------------------------------------------------------------------
# Health
# -----------------------------------------------------------------------------
@app.get("/healthz")
def healthz():
    return {"ok": True}

# -----------------------------------------------------------------------------
# Protected: XLSX → ZIP
# -----------------------------------------------------------------------------
@app.post("/api/upload")
async def upload_excel(file: UploadFile = File(...), email: str = Depends(require_user)):
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx file.")

    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="Empty file.")

    try:
        wb = load_workbook(io.BytesIO(data), data_only=False, read_only=False, keep_links=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Unable to read the Excel file. Is it valid .xlsx?")

    workbook_name = os.path.splitext(os.path.basename(file.filename))[0]
    zip_buf = io.BytesIO()
    failed: List[Dict[str, Any]] = []

    with zipfile.ZipFile(zip_buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        async with httpx.AsyncClient(follow_redirects=True, headers=DEFAULT_HEADERS, timeout=45.0) as client:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                loc = find_final_col(ws)
                if not loc:
                    continue
                header_row, col_idx = loc

                plans: List[Tuple[str, List[str]]] = []
                for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                    raw = extract_url_from_cell(row[0])
                    if not raw:
                        continue
                    display, candidates = build_fetch_plan(raw)
                    if candidates:
                        plans.append((display, candidates))
                if not plans:
                    continue

                inner_dir = f"{slugify(workbook_name)}/{slugify(sheet_name)}/images"
                sem = asyncio.Semaphore(10)
                results = await asyncio.gather(*(fetch_one(client, cand, sem) for _, cand in plans))

                seq = 1
                for (display_url, _), content in zip(plans, results):
                    if not content:
                        failed.append({"sheet": sheet_name, "url": display_url})
                        continue             # Do NOT increment seq
                    try:
                        img = Image.open(io.BytesIO(content))
                        if img.mode not in ("RGB", "L"):
                            img = img.convert("RGB")
                        out = io.BytesIO()
                        img.save(out, format="JPEG", quality=95)
                        name = f"final_image_{seq}.jpg"
                        zf.writestr(f"{inner_dir}/{name}", out.getvalue())
                        seq += 1             # Only increment here on success
                    except (UnidentifiedImageError, Exception):
                        failed.append({"sheet": sheet_name, "url": display_url})
                        continue             # Do NOT increment seq


        if failed:
            zf.writestr(f"{slugify(workbook_name)}/failed.json", json.dumps(failed, ensure_ascii=False, indent=2))

    zip_buf.seek(0)
    original_zip = f"{workbook_name}_images.zip"
    ascii_name = ascii_fallback_filename(original_zip)
    filename_star = quote(original_zip, safe="")
    failed_b64 = base64.b64encode(json.dumps(failed).encode("utf-8")).decode("ascii")
    cd = f'attachment; filename="{ascii_name}"; filename*=UTF-8\'\'{filename_star}'

    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={"Content-Disposition": cd, "X-Failed-Json": failed_b64},
    )

# -----------------------------------------------------------------------------
# Static UI (served by FastAPI) - mount AFTER API routes
# -----------------------------------------------------------------------------
if os.path.isdir("static"):
    app.mount("/", StaticFiles(directory="static", html=True), name="static")
